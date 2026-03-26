"""
合规检查工作底稿自动生成工具
主入口函数
"""
import datetime
import re
from pathlib import Path
from typing import List, Optional, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# 使用本项目内部的模块
from src.utils.database_factory import DatabaseFactory
from src.utils.logger import logger
from src.utils.models import ModelFactory, stream_wrapper

# 导入规划模块
from src.planner_modules.modules import parse_file, chunk_regulation_text, formulate_method, take_shots

from langchain_core.messages import SystemMessage, HumanMessage, AIMessage
# Monkey patch for reasoning content
# ==========================================================================
from langchain_openai.chat_models import base as langchain_openai_base
_original_convert_delta = langchain_openai_base._convert_delta_to_message_chunk
def _patched_convert_delta(_dict, default_class):
    chunk = _original_convert_delta(_dict, default_class)
    if "reasoning_content" in _dict:
        chunk.additional_kwargs["reasoning_content"] = _dict["reasoning_content"]
    return chunk
langchain_openai_base._convert_delta_to_message_chunk = _patched_convert_delta
# ==========================================================================

# 5大类材料定义
MATERIAL_CATEGORIES = {
    "规章制度与规范文件类": [
        "公司章程、有效制度清单",
        "廉洁从业管理实施细则 / 廉洁从业规范",
        "各业务条线管理制度",
        "财务费用管理办法",
        "采购管理相关制度 / 招投标管理办法",
        "保密管理规定 / 信息隔离墙制度",
        "居间人管理办法 / 投资顾问管理制度",
    ],
    "底层数据与台账明细类": [
        "费用支出明细",
        "员工/客户 MAC/IP 地址统计表",
        "客户保证金调整名单及明细",
        "员工及近亲属投资账户申报台账",
        "信访举报登记台账",
        "公司资管产品交易流水 / 自营账户交易流水",
    ],
    "审批痕迹与会议纪要类": [
        "董事会决议 / 监事会工作报告",
        "投资决策委员会会议纪要",
        "合规内控会议纪要",
        "OA系统内部决策流程审批单 / 费用报销单",
        "数据导出/系统最高权限访问审批单",
    ],
    "人事与合规监督档案类": [
        "员工签署的《廉洁从业承诺书》",
        "员工培训记录档案",
        "员工电脑登录情况排查报告",
        "历史年度人员离任审计报告",
        "纪律处分/内部问责通报文件",
    ],
    "第三方合作与业务材料类": [
        "法律顾问名单及准入材料",
        "居间人名单 / 投资顾问清单",
        "居间协议 / 投顾协议 / 采购合同",
        "研究报告",
        "供应商比价单 / 招投标记录文件",
    ],
}


def backfill_database_embeddings(schema_name: str = 'compliance_review', table_name: str = 'shot_table'):
    """
    数据库向量补全函数
    
    使用工厂模式延迟获取数据库连接。
    
    Args:
        schema_name: 数据库 schema 名称
        table_name: 表名称
    """
    from src.shot_table_modules.modules import backfill_embeddings
    return backfill_embeddings(schema_name, table_name)


def generate_audit_draft(
    file_paths: List[Union[str, Path]],
    output_path: Optional[Union[str, Path]] = None
) -> Path:
    """
    生成检查工作底稿Excel文件
    
    Args:
        file_paths: 合规文件路径列表（支持多份文件）
        output_path: 可选的输出路径
    
    Returns:
        生成的Excel文件路径
    """
    file_paths = [Path(p) for p in file_paths]
    
    # 1. 解析所有文件
    logger.info(f"[Step 1] 解析 {len(file_paths)} 份文件...")
    all_articles = []
    
    for file_path in file_paths:
        doc_name = file_path.stem
        logger.info(f"  - 解析: {doc_name}")
        
        # 使用 MinerU 解析文件
        raw_content = parse_file(file_path, is_ocr=True)
        
        # 提取条款
        articles = chunk_regulation_text(raw_content, doc_name=doc_name)
        
        all_articles.extend(articles)
        logger.info(f"    提取 {len(articles)} 条条款")
    
    logger.info(f"  总计: {len(all_articles)} 条条款")

    # 2. 拆分长条款
    logger.info(f"[Step 2] 拆分长条款...")
    final_articles = []
    for article in all_articles:
        # 将条款转换成检查内容的函数，某种程度上的洗稿或者拆解
        split_items = _split_and_reform_article(article)
        final_articles.extend(split_items)

    logger.info(f"  拆分后: {len(final_articles)} 条检查内容")

    # 3. 使用工厂模式获取数据库连接和模型（延迟初始化）
    postgresql_server = DatabaseFactory.get_postgresql_server()
    rag_server = DatabaseFactory.get_rag_server()

    # 4. 生成检查内容
    logger.info(f"[Step 3] 生成检查内容...")
    audit_items = []

    for idx, article in enumerate(final_articles, 1):
        logger.info(f"  - 进度: {idx}/{len(final_articles)}")

        # 原始 content = 检查依据（来自哪些条款）
        audit_content = article['content']
        sources = article.get('metadata', {})

        # 获取 shots（RAG 样例）
        shots_gen = take_shots(
            postgresql_server,
            rag_server,
            [article],
            'compliance_review',
            'shot_table'
        )
        shots = []
        # 这个地方本来如果传入的列表中有多个对象的话一定是错的
        # 但是因为这个地方列表里永远只会有一个对象，负负得正了
        for _, shot_list in shots_gen:
            shots = shot_list
            break

        # 生成检查方式
        check_method_dic = formulate_method(audit_content, shots)

        # 如果返回的是 dict，提取 response 字段
        if isinstance(check_method_dic, dict):
            check_method = check_method_dic.get('response', str(check_method_dic))
            check_files = check_method_dic.get('files','')
        else:
            check_method = str(check_method_dic)
            check_files = ''

        # 格式化检查依据（引用原始条款 A）
        audit_basis = _format_sources(sources)

        audit_items.append({
            '序号': idx,
            '检查内容': audit_content,
            '检查依据': audit_basis,
            '检查方式': check_method,
            '对应材料': check_files,
            '检查发现问题': '',
            '备注': ''
        })
    
    # 5. 导出Excel
    logger.info(f"[Step 4] 导出Excel文件...")
    if output_path is None:
        output_dir = Path(__file__).parent.parent / 'results'
        output_dir.mkdir(exist_ok=True)
        output_path = output_dir / f'检查工作底稿_{datetime.datetime.now().strftime("%m%d_%H%M%S")}.xlsx'
    
    _export_to_excel(audit_items, output_path)
    logger.info(f"  输出: {output_path}")
    
    return output_path


def _split_and_reform_article(article: dict) -> List[dict]:
    """
    使用 LLM 将长条款拆分为多个可独立执行的检查条目。
    拆分依据是"可操作性"，而非原始分点数量。
    拆分后的内容不得与原条款完全相同。
    即使短的法条也要洗稿

    Args:
        article: 包含 content 和 metadata 的条款

    Returns:
        拆分后的条款列表
    """
    content = article.get('content', '')
    metadata = article.get('metadata', {})

    model = ModelFactory.get_method_model()

    if len(content)>=200:
        system_prompt = f"""
<role>
你是一位资深金融合规审查专家。
</role>

<task>
现有的合规条款太冗长或者太空虚，请提取出其中可以实操的"检查内容"。请按照步骤完成任务
## 步骤
1. 提取条款原文中针对检查方法的内容，避免任何宏大叙事，处罚规则和无法实操、无法检查的内容
2. 将第 1 步中的内容基于不同侧重点进行切分，每一块内容应该通顺，逻辑自洽，尊重原文，遵从最小修改原则。若条款原文已经有分点，可以直接使用，若没有分点，请自行切分总结，通常切为2~4块
3. 将第 2 步中的内容整理为每块附带前缀"###"并进行输出
</task>

<output_format>
## 输出要求
直接输出拆分结果，每条一行，用 "###" 分隔开头，不要任何解释或序号前缀。
</output_format>
"""
        example_input = HumanMessage(
            content='第四条  期货经营机构承担廉洁从业风险防控主体责任，应当积极推进新时代廉洁文化建设，并在公司重要制度中以专门章节明确廉洁从业要求，建立涵盖所有业务及各个环节的廉洁从业内部控制制度，将其纳入整个内部控制体系之中，制定具体、有效的事前风险防范体系、事中管控措施和事后追责机制，明确董事、监事、高级管理人员及各层级管理人员的廉洁从业管理职责，主动防范、应对、报告廉洁从业风险，并对廉洁从业风险管控工作的相关底稿留档保存。'
        )
        example_output = AIMessage(
            content='###应当在公司重要制度中以专门章节明确廉洁从业要求，建立涵盖所有业务及各个环节的廉洁从业内部控制制度，将其纳入整个内部控制体系之中\n###明确董事、监事、高级管理人员及各层级管理人员的廉洁从业管理职责\n###对廉洁从业风险管控工作的相关底稿留档保存'
        )
    else:
        system_prompt = f"""
<role>
你是一位资深金融合规审查专家。
</role>

<task>
请通过同义词替换，扩句等手段将文档法规进行改写，要求保持含义基本一致，但是文字描述不同
对于同一条法规，你只需输出**一条**替换后的结果即可
</task>

<output_format>
## 输出要求
直接输出拆分结果，以"###"开头，不要任何解释或序号前缀。
</output_format>
        """
        example_input = HumanMessage(
            content='第十一条  期货经营机构应当指定与业务职责不相冲突的部门或者岗位对本机构及其工作人员的廉洁从业情况进行监督，定期或者不定期开展廉洁从业内部检查，对发现的问题及时整改，对责任人按照有关规定严肃处理。责任人为中国共产党党员的，同时按照党的纪律要求进行处理。'
        )
        example_output = AIMessage(
            content='###期货经营机构应当指定部门对本机构及其工作人员的廉洁从业情况进行监督，定期或不定期开展廉洁从业内部检查'
        )

    try:
        system_msg = SystemMessage(
            content=system_prompt
        )
        human_msg = HumanMessage(
            content=content
        )
        context = [system_msg,example_input,example_output,human_msg]

        response = stream_wrapper(model, context)
        reform_result = response.content.strip()
        logger.info(f"[总结] LLM 总结完成")

        # 解析拆分结果：按 "###" 分割
        raw_items = re.split(r'\n###\s*', reform_result)
        # 如果没有 "###" 前缀，直接按行分割
        if len(raw_items) <= 1:
            raw_items = [line.strip() for line in reform_result.split('\n') if line.strip()]

        split_items = []
        for item_text in raw_items:
            item_text = item_text.strip()
            # 过滤掉空的、或者和原文一模一样的
            if not item_text:
                continue
            # 过滤掉特别短的（可能是解释性文字残留）
            if len(item_text) < 8:
                continue
            split_items.append({
                'content': item_text.removeprefix('###'),
                'metadata': metadata
            })

        # 如果解析不出有效条目，回退到原文
        if not split_items:
            logger.warning("[拆分] LLM 拆分结果为空，回退为原文")
            return [article]

        return split_items

    except Exception as e:
        logger.warning(f"[拆分] LLM 调用失败，回退为原文: {e}")
        return [article]


def _format_sources(sources: dict) -> str:
    """格式化检查依据"""
    doc_name = sources.get('doc_name', '')
    article_num = sources.get('article_num', '')
    if doc_name and article_num:
        logger.info(f'获取到文件名和条款编号，拼接为检查依据')
        source_result = f"《{doc_name}》{article_num}"
    elif not doc_name:
        logger.error(f'未能获取到文件名，无法输出完整检查依据')
        source_result = f'《未知文件》{article_num}'
    else:
        logger.error(f'未能获取到条款编号，无法输出完整检查依据')
        source_result = f'《{doc_name}》'
    
    return source_result


def _select_materials(content: str) -> str:
    """选择对应材料（关键词匹配）"""
    materials = []
    content_lower = content.lower()
    
    # 关键词匹配
    keywords_map = {
        '培训': '员工培训记录档案',
        '承诺': '员工签署的《廉洁从业承诺书》',
        '制度': '公司章程、有效制度清单',
        '费用': '费用支出明细',
        '招待': '费用支出明细（招待费）',
        '居间': '居间人名单 / 居间协议',
        '投顾': '投资顾问清单 / 投顾协议',
        '交易': '公司资管产品交易流水 / 自营账户交易流水',
        '会议': '合规内控会议纪要',
        '审批': 'OA系统内部决策流程审批单',
        '采购': '采购管理相关制度 / 招投标管理办法',
        '保密': '保密管理规定 / 信息隔离墙制度',
        '举报': '信访举报登记台账',
        '审计': '历史年度人员离任审计报告',
    }
    
    for keyword, material in keywords_map.items():
        if keyword in content_lower:
            if material not in materials:
                materials.append(material)
    
    return "；".join(materials[:3]) if materials else "相关制度文件；台账记录"


def _export_to_excel(
    audit_items: List[dict],
    output_path: Path
):
    """导出到Excel"""
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = '检查工作底稿'
    
    # 写入表头
    ws.merge_cells('A1:G1')
    ws.cell(row=1, column=1, value='合规检查工作底稿')
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    # 基本信息
    ws.cell(row=2, column=3, value='检查项目')
    ws.cell(row=2, column=5, value='检查时间')
    ws.cell(row=3, column=3, value='被检查单位')
    ws.cell(row=3, column=5, value='检查主办部门')
    ws.cell(row=4, column=3, value='检查人员')
    
    # 列标题
    headers = ['序号', '检查内容', '检查依据', '检查方式', '对应材料', '检查发现问题', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # 写入数据
    for idx, item in enumerate(audit_items, 7):
        ws.cell(row=idx, column=1, value=item.get('序号', ''))
        ws.cell(row=idx, column=2, value=item.get('检查内容', ''))
        ws.cell(row=idx, column=3, value=item.get('检查依据', ''))
        ws.cell(row=idx, column=4, value=item.get('检查方式', ''))
        ws.cell(row=idx, column=5, value=item.get('对应材料', ''))
        ws.cell(row=idx, column=6, value=item.get('检查发现问题', ''))
        ws.cell(row=idx, column=7, value=item.get('备注', ''))
    
    # 设置样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    column_widths = {'A': 8, 'B': 50, 'C': 30, 'D': 60, 'E': 40, 'F': 20, 'G': 15}
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    for row in ws.iter_rows(min_row=6, max_row=6+len(audit_items), min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # 保存
    wb.save(output_path)


if __name__ == '__main__':
    files = [
        r"C:\Users\65180\Desktop\Pengbo\current_materials\廉洁检查材料（测试用）\廉洁检查材料（测试用）\内外部规定\1.证券期货经营机构及其工作人员廉洁从业规定.docx",
        r"C:\Users\65180\Desktop\Pengbo\current_materials\廉洁检查材料（测试用）\廉洁检查材料（测试用）\内外部规定\2.期货经营机构及其工作人员廉洁从业实施细则.docx"
    ]
    result = generate_audit_draft(files)
    print(f"生成完成: {result}")